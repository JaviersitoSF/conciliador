import csv
import os
import platform
import subprocess
import sys
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import pandas as pd
from num2words import num2words
import xlrd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

ARCHIVO_CHEQUES = "cheques_emitidos.csv"
ARCHIVO_BANCO = "estado_cuenta.xlsx"
ARCHIVO_BANCO_CSV = "estado_cuenta.csv"
ARCHIVO_DEPOSITOS = "depositos.csv"
TIPO_CHEQUE = "CHEQUE"
TIPO_NOTA_DEBITO = "NOTA_DEBITO"
ESTADO_TRANSITO = "PENDIENTE"
ESTADO_RECONCILIADO = "RECONCILIADO"
ESTADO_DIFERENCIA = "DIFERENCIA"
ESTADO_INVALIDO = "INVALIDO"
ESTADO_ANULADO = "ANULADO"
ESTADO_CONFLICTO = "CONFLICTO"


class ErrorOperacion(ValueError):
    """Error esperado en operaciones del sistema."""


def crear_dataframe_vacio(columnas):
    df = pd.DataFrame(columns=columnas + ["Monto_valor", "Fecha_dt"])
    df["Monto_valor"] = pd.Series(dtype=object)
    df["Fecha_dt"] = pd.to_datetime(pd.Series(dtype=object))
    return df


def convertir_monto(valor):
    if valor is None:
        return None

    if isinstance(valor, Decimal):
        return valor.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    try:
        if pd.isna(valor):
            return None
    except TypeError:
        pass

    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return None

    negativo = texto.startswith("(") and texto.endswith(")")
    if negativo:
        texto = texto[1:-1]

    texto = (
        texto.replace("Q", "")
        .replace("q", "")
        .replace("$", "")
        .replace("\xa0", "")
        .replace(" ", "")
        .replace(",", "")
        .replace("−", "-")
    )

    if texto in {"", "-", ".", "-.", "--"}:
        return None

    try:
        monto = Decimal(texto)
    except InvalidOperation:
        return None

    if negativo:
        monto = -monto

    return monto.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def formatear_monto(valor):
    monto = convertir_monto(valor)
    if monto is None:
        raise ValueError("Monto invalido")
    return f"{monto:.2f}"


def normalizar_texto_clave(valor):
    texto = str(valor or "").strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
    return " ".join(texto.split())


def normalizar_numero_cheque(valor):
    if valor is None:
        return None

    try:
        if pd.isna(valor):
            return None
    except TypeError:
        pass

    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return None

    if texto.endswith(".0"):
        texto = texto[:-2]

    if texto.isdigit():
        numero = int(texto)
        return str(numero) if numero > 0 else None

    try:
        numero_decimal = Decimal(texto)
    except InvalidOperation:
        return None

    if numero_decimal != numero_decimal.to_integral_value(rounding=ROUND_HALF_UP):
        return None

    numero = int(numero_decimal)
    return str(numero) if numero > 0 else None


def normalizar_referencia(valor):
    texto = normalizar_texto_clave(valor)
    return texto or None


def convertir_fecha_flexible(valor):
    if valor is None:
        return pd.NaT

    try:
        if pd.isna(valor):
            return pd.NaT
    except TypeError:
        pass

    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return pd.NaT

    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return pd.to_datetime(texto, format=formato)
        except Exception:
            continue

    fecha = pd.to_datetime(texto, errors="coerce", dayfirst=True)
    if pd.notna(fecha):
        return fecha

    return pd.to_datetime(texto, errors="coerce")


def pedir_texto_no_vacio(mensaje):
    while True:
        texto = input(mensaje).strip()
        if texto:
            return texto
        print("⚠️ Error: el campo no puede quedar vacío.")


def pedir_monto_positivo(mensaje):
    while True:
        texto = input(mensaje).strip()
        monto = convertir_monto(texto)
        if monto is None:
            print("⚠️ Error: Solo usar números y punto decimal.")
            continue
        if monto <= 0:
            print("⚠️ Error: El monto debe ser mayor que cero.")
            continue
        return monto


def pedir_numero_cheque(mensaje):
    while True:
        texto = input(mensaje).strip()
        if not texto:
            print("⚠️ Error: el número de cheque no puede quedar vacío.")
            continue
        if not texto.isdigit():
            print("⚠️ Error: el número de cheque debe contener solo dígitos.")
            continue

        numero = int(texto)
        if numero <= 0:
            print("⚠️ Error: el número de cheque debe ser mayor que cero.")
            continue

        return str(numero)


def leer_csv_flexible(ruta, encoding="utf-8"):
    filas = []
    with open(ruta, newline="", encoding=encoding) as archivo:
        lector = csv.reader(archivo)
        for fila in lector:
            filas.append(fila)

    if not filas:
        return pd.DataFrame()

    max_columnas = max(len(fila) for fila in filas)
    filas_normalizadas = [fila + [""] * (max_columnas - len(fila)) for fila in filas]
    return pd.DataFrame(filas_normalizadas)


def cargar_cheques_registrados():
    columnas_base = ["Num", "Fecha", "Nombre", "Monto", "Estado", "Tipo", "Conciliacion"]
    vacio = crear_dataframe_vacio(columnas_base[:6])
    vacio["Num_norm"] = pd.Series(dtype=object)
    vacio["Conciliacion"] = pd.Series(dtype=object)

    if not os.path.exists(ARCHIVO_CHEQUES):
        return vacio

    try:
        df = leer_csv_flexible(ARCHIVO_CHEQUES)
    except FileNotFoundError:
        return vacio
    except Exception:
        return vacio

    if df.empty:
        return vacio

    if df.shape[1] >= 7:
        nombres = columnas_base
    elif df.shape[1] == 6:
        nombres = columnas_base[:6]
    else:
        nombres = columnas_base[:5]

    while df.shape[1] < len(nombres):
        df[df.shape[1]] = ""
    df = df.iloc[:, : len(nombres)].copy()
    df.columns = nombres

    df = df.fillna("")
    df = df[df[nombres].apply(lambda fila: any(str(valor).strip() for valor in fila), axis=1)].copy()

    for columna in columnas_base:
        if columna not in df.columns:
            df[columna] = ""

    for columna in columnas_base:
        df[columna] = df[columna].astype(str).str.strip()

    df["Estado"] = df["Estado"].str.upper()
    df.loc[df["Estado"].isin(["", "NAN"]), "Estado"] = "TRANSITO"
    df["Tipo"] = df["Tipo"].str.upper()
    df.loc[df["Tipo"].isin(["", "NAN"]), "Tipo"] = TIPO_CHEQUE
    df["Conciliacion"] = df["Conciliacion"].str.upper()
    df.loc[df["Conciliacion"].isin(["", "NAN"]), "Conciliacion"] = ESTADO_TRANSITO
    df["Num_norm"] = df["Num"].map(normalizar_numero_cheque)
    df["Monto_valor"] = df["Monto"].map(convertir_monto)
    df["Fecha_dt"] = df["Fecha"].map(convertir_fecha_flexible)

    return df


def cargar_depositos_registrados():
    columnas_base = ["Fecha", "Descripcion", "Monto", "Conciliacion"]
    vacio = crear_dataframe_vacio(columnas_base[:3])
    vacio["Conciliacion"] = pd.Series(dtype=object)

    if not os.path.exists(ARCHIVO_DEPOSITOS):
        return vacio

    try:
        df = leer_csv_flexible(ARCHIVO_DEPOSITOS)
    except FileNotFoundError:
        return vacio
    except Exception:
        return vacio

    if df.empty:
        return vacio

    if df.shape[1] >= 4:
        nombres = columnas_base
    else:
        nombres = columnas_base[:3]

    while df.shape[1] < len(nombres):
        df[df.shape[1]] = ""
    df = df.iloc[:, : len(nombres)].copy()
    df.columns = nombres

    df = df.fillna("")
    df = df[df[nombres].apply(lambda fila: any(str(valor).strip() for valor in fila), axis=1)].copy()

    for columna in columnas_base:
        if columna not in df.columns:
            df[columna] = ""

    for columna in columnas_base:
        df[columna] = df[columna].astype(str).str.strip()

    df["Descripcion"] = df["Descripcion"].str.upper()
    df["Conciliacion"] = df["Conciliacion"].str.upper()
    df.loc[df["Conciliacion"].isin(["", "NAN"]), "Conciliacion"] = ESTADO_TRANSITO
    df["Monto_valor"] = df["Monto"].map(convertir_monto)
    df["Fecha_dt"] = df["Fecha"].map(convertir_fecha_flexible)

    return df


def buscar_archivo_estado_cuenta():
    if os.path.exists("Transacciones del mes.xls"):
        return "Transacciones del mes.xls"

    if os.path.exists("Transacciones del mes.xlsx"):
        return "Transacciones del mes.xlsx"

    if os.path.exists(ARCHIVO_BANCO):
        return ARCHIVO_BANCO

    if os.path.exists(ARCHIVO_BANCO_CSV):
        return ARCHIVO_BANCO_CSV

    csvs_bancarios = [
        archivo
        for archivo in os.listdir(".")
        if archivo.lower().endswith(".csv")
        and archivo not in {ARCHIVO_CHEQUES, ARCHIVO_DEPOSITOS}
    ]

    if len(csvs_bancarios) == 1:
        return csvs_bancarios[0]

    if len(csvs_bancarios) > 1:
        raise ErrorOperacion(
            "⚠️ Hay varios CSV externos. Deja solo el estado de cuenta o nómbralo estado_cuenta.csv."
        )

    return None


def detectar_fila_encabezado_csv(ruta):
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(ruta, newline="", encoding=encoding) as archivo:
                lector = csv.reader(archivo)
                for indice, fila in enumerate(lector):
                    if fila_parece_encabezado_banco(fila):
                        return indice, encoding
        except UnicodeDecodeError:
            continue

    raise ErrorOperacion("⚠️ No pude encontrar el encabezado del estado de cuenta CSV.")


def fila_parece_encabezado_banco(valores):
    columnas = {normalizar_texto_clave(columna) for columna in valores if str(columna or "").strip()}
    if not columnas:
        return False

    if {"NUM_CHEQUE", "MONTO"}.issubset(columnas):
        return True

    if {"DEBE", "HABER"}.issubset(columnas):
        return True

    if "FECHA" in columnas:
        candidatas = {
            "TT",
            "DESCRIPCION",
            "NO. DOC",
            "NO DOC",
            "DOCUMENTO",
            "REFERENCIA",
            "MONTO",
            "DEBE",
            "HABER",
            "SALDO",
        }
        if columnas.intersection(candidatas):
            return True

    return False


def cargar_estado_cuenta_csv(ruta):
    fila_encabezado, encoding = detectar_fila_encabezado_csv(ruta)
    try:
        return pd.read_csv(
            ruta,
            skiprows=fila_encabezado,
            header=0,
            dtype=str,
            keep_default_na=False,
            encoding=encoding,
        )
    except Exception as e:
        raise ErrorOperacion(f"⚠️ No pude leer el CSV del banco: {e}") from e


def detectar_hoja_y_encabezado_excel(ruta):
    extension = os.path.splitext(ruta)[1].lower()

    if extension == ".xls":
        try:
            workbook = xlrd.open_workbook(ruta)
        except Exception as e:
            raise ErrorOperacion(f"⚠️ No pude abrir el estado de cuenta Excel: {e}") from e

        for hoja in workbook.sheets():
            for indice in range(hoja.nrows):
                fila = hoja.row_values(indice)
                if fila_parece_encabezado_banco(fila):
                    return hoja.name, indice + 1
    else:
        try:
            workbook = load_workbook(ruta, read_only=True, data_only=True)
        except Exception as e:
            raise ErrorOperacion(f"⚠️ No pude abrir el estado de cuenta Excel: {e}") from e

        for hoja in workbook.worksheets:
            for indice, fila in enumerate(hoja.iter_rows(values_only=True), start=1):
                if fila_parece_encabezado_banco(fila):
                    return hoja.title, indice

    raise ErrorOperacion(
        "⚠️ El archivo del banco debe incluir un encabezado reconocible como "
        "'Fecha', 'TT', 'Descripción', 'No. Doc', 'Debe', 'Haber' o 'Num_cheque' y 'Monto'."
    )


def cargar_estado_cuenta_excel(ruta):
    hoja, fila_encabezado = detectar_hoja_y_encabezado_excel(ruta)
    try:
        extension = os.path.splitext(ruta)[1].lower()
        return pd.read_excel(
            ruta,
            sheet_name=hoja,
            skiprows=fila_encabezado - 1,
            header=0,
            dtype=str,
            keep_default_na=False,
            engine="xlrd" if extension == ".xls" else None,
        )
    except Exception as e:
        raise ErrorOperacion(f"⚠️ Ocurrió un error leyendo el estado de cuenta Excel: {e}") from e


def obtener_columna(columnas, *candidatas):
    for candidata in candidatas:
        columna = columnas.get(normalizar_texto_clave(candidata))
        if columna:
            return columna
    return None


def cargar_estado_cuenta_banco():
    ruta = buscar_archivo_estado_cuenta()
    if not ruta:
        raise ErrorOperacion("⚠️ Faltan archivos. Asegúrate de tener registros y estado de cuenta.")

    extension = os.path.splitext(ruta)[1].lower()
    if extension in {".csv", ".txt"}:
        df_banco = cargar_estado_cuenta_csv(ruta)
    elif extension in {".xlsx", ".xlsm", ".xls"}:
        df_banco = cargar_estado_cuenta_excel(ruta)
    else:
        try:
            df_banco = pd.read_excel(ruta)
        except Exception as e:
            raise ErrorOperacion(f"⚠️ Ocurrió un error leyendo el estado de cuenta: {e}") from e

    if df_banco.empty:
        return crear_dataframe_vacio(
            ["Fecha", "TT", "Descripcion", "Referencia", "Debe", "Haber", "Saldo", "Tipo_movimiento"]
        )

    columnas = {normalizar_texto_clave(columna): columna for columna in df_banco.columns}
    col_fecha = obtener_columna(columnas, "Fecha", "Fecha operación", "Fecha contable")
    col_tt = obtener_columna(columnas, "TT", "Tipo", "Tipo Transacción")
    col_descripcion = obtener_columna(columnas, "Descripción", "Descripcion", "Concepto", "Detalle")
    col_referencia = obtener_columna(columnas, "No. Doc", "No Doc", "Documento", "Referencia", "Num_cheque")
    col_debe = obtener_columna(
        columnas,
        "Debe (GTQ)",
        "Debe",
        "Débito",
        "Debito",
        "Débitos",
        "Debitos",
        "Retiros",
        "Cargos",
    )
    col_haber = obtener_columna(
        columnas,
        "Haber (GTQ)",
        "Haber",
        "Crédito",
        "Credito",
        "Créditos",
        "Creditos",
        "Depósitos",
        "Depositos",
        "Abonos",
    )
    col_monto = obtener_columna(columnas, "Monto", "Importe", "Valor")
    col_saldo = obtener_columna(columnas, "Saldo (GTQ)", "Saldo")

    if not col_monto and not (col_debe or col_haber):
        raise ErrorOperacion(
            "⚠️ El archivo del banco debe incluir columnas de monto, debe/haber o cargos/abonos."
        )

    df = df_banco.copy().fillna("")
    normalizado = pd.DataFrame()
    normalizado["Fecha"] = df[col_fecha].astype(str).str.strip() if col_fecha else ""
    normalizado["TT"] = df[col_tt].astype(str).str.strip().str.upper() if col_tt else ""
    normalizado["Descripcion"] = df[col_descripcion].astype(str).str.strip() if col_descripcion else ""
    normalizado["Referencia"] = df[col_referencia].astype(str).str.strip() if col_referencia else ""
    normalizado["Debe"] = df[col_debe].map(convertir_monto) if col_debe else None
    normalizado["Haber"] = df[col_haber].map(convertir_monto) if col_haber else None
    normalizado["Saldo"] = df[col_saldo].map(convertir_monto) if col_saldo else None

    if col_monto:
        montos = df[col_monto].map(convertir_monto)
        hay_montos_negativos = any(monto is not None and monto < 0 for monto in montos)
        normalizado["Monto_valor"] = montos.map(lambda monto: abs(monto) if monto is not None else None)
        normalizado["Tipo_movimiento"] = montos.map(
            lambda monto: "CREDITO" if hay_montos_negativos and monto is not None and monto > 0 else "DEBITO"
        )
    else:
        def elegir_monto(fila):
            debe = fila["Debe"]
            haber = fila["Haber"]
            if debe is not None and debe > 0:
                return debe
            if haber is not None and haber > 0:
                return haber
            return debe if debe is not None else haber

        def elegir_tipo(fila):
            debe = fila["Debe"]
            haber = fila["Haber"]
            if debe is not None and debe > 0:
                return "DEBITO"
            if haber is not None and haber > 0:
                return "CREDITO"
            return "DEBITO"

        normalizado["Monto_valor"] = normalizado.apply(elegir_monto, axis=1)
        normalizado["Tipo_movimiento"] = normalizado.apply(elegir_tipo, axis=1)

    normalizado["TT"] = normalizado["TT"].astype(str).str.upper()
    normalizado.loc[normalizado["TT"].isin(["CQ", "ND"]), "Tipo_movimiento"] = "DEBITO"
    normalizado.loc[normalizado["TT"].isin(["DE", "NC"]), "Tipo_movimiento"] = "CREDITO"
    normalizado["Fecha_dt"] = normalizado["Fecha"].map(convertir_fecha_flexible)
    normalizado["Num_norm"] = normalizado["Referencia"].map(normalizar_numero_cheque)
    normalizado["Referencia_norm"] = normalizado["Referencia"].map(normalizar_referencia)
    normalizado["Descripcion_norm"] = normalizado["Descripcion"].map(normalizar_texto_clave)
    normalizado["TT"] = normalizado["TT"].where(normalizado["TT"].astype(str).str.strip() != "", normalizado["Tipo_movimiento"])
    normalizado = normalizado[
        normalizado["Monto_valor"].notna()
        | normalizado["Num_norm"].notna()
        | normalizado["Referencia_norm"].notna()
    ].copy()

    return normalizado


def cheque_ya_registrado(numero):
    df = cargar_cheques_registrados()
    if df.empty or "Num_norm" not in df.columns:
        return False
    return df["Num_norm"].eq(numero).any()


def nota_debito_ya_registrada(referencia):
    referencia = str(referencia or "").strip().upper()
    if not referencia:
        return False

    df = cargar_cheques_registrados()
    if df.empty or "Num" not in df.columns or "Tipo" not in df.columns:
        return False

    referencias = df["Num"].astype(str).str.strip().str.upper()
    tipos = df["Tipo"].astype(str).str.strip().str.upper()
    return (referencias.eq(referencia) & tipos.eq(TIPO_NOTA_DEBITO)).any()


def registrar_deposito_datos(monto, descripcion, fecha=None):
    monto = convertir_monto(monto)
    if monto is None:
        raise ErrorOperacion("⚠️ Error: Solo usar números y punto decimal.")
    if monto <= 0:
        raise ErrorOperacion("⚠️ Error: El monto debe ser mayor que cero.")

    descripcion = str(descripcion or "").strip()
    if not descripcion:
        raise ErrorOperacion("⚠️ Error: el campo no puede quedar vacío.")

    fecha = fecha or datetime.now().strftime("%Y-%m-%d")
    descripcion = descripcion.upper()
    conciliacion = ESTADO_TRANSITO

    with open(ARCHIVO_DEPOSITOS, mode="a", newline="", encoding="utf-8") as f:
        escritor = csv.writer(f)
        escritor.writerow([fecha, descripcion, formatear_monto(monto), conciliacion])

    return {
        "fecha": fecha,
        "descripcion": descripcion,
        "monto": monto,
        "mensaje": f"✅ Depósito de Q {formatear_monto(monto)} registrado con éxito.",
    }


def registrar_deposito():
    print("\n--- REGISTRO DE DEPÓSITO ---")
    monto = pedir_monto_positivo("Monto del depósito (ej. 5000.00): ")
    descripcion = pedir_texto_no_vacio("Descripción (ej. Venta mostrador, Eukanuba): ").upper()
    resultado = registrar_deposito_datos(monto, descripcion)
    print(resultado["mensaje"])


def imprimir_cheque_pdf(num, fecha, nombre, monto):
    ancho_papel = 21.5 * cm
    alto_papel = 14.0 * cm
    nombre_pdf = f"cheque_{num}.pdf"

    pdf = canvas.Canvas(nombre_pdf, pagesize=(ancho_papel, alto_papel))

    monto_formateado = formatear_monto(monto)
    entero, centavos = monto_formateado.split(".")
    monto_en_letras = num2words(int(entero), lang="es").upper()
    texto_oficial = f"La suma de: {monto_en_letras} QUETZALES CON {centavos}/100"

    pdf.drawString(15 * cm, 12 * cm, f"Fecha: {fecha}")
    pdf.drawString(2 * cm, 11 * cm, f"Páguese a: {nombre}")
    pdf.drawString(16 * cm, 11 * cm, f"Q {monto_formateado}")
    pdf.drawString(2 * cm, 10 * cm, texto_oficial)

    pdf.save()

    print(f"🖨️  PDF generado: {nombre_pdf}")

    try:
        if platform.system() == "Windows":
            startfile = getattr(os, "startfile", None)
            if callable(startfile):
                startfile(nombre_pdf, "print")
            else:
                raise AttributeError("os.startfile no esta disponible en este entorno")
        elif platform.system() == "Darwin":
            abrir_pdf_silenciosamente(["open", nombre_pdf])
        else:
            print("🐧 Abriendo en Chrome OS...")
            abrir_pdf_silenciosamente(["xdg-open", nombre_pdf])
    except Exception as e:
        print(f"⚠️ Abre el PDF manual. Error: {e}")


def abrir_pdf_silenciosamente(comando):
    try:
        subprocess.run(
            comando,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True,
        )
    except subprocess.CalledProcessError as e:
        detalle = (e.stderr or e.stdout or "").strip()
        if detalle:
            raise RuntimeError(detalle) from e
        raise RuntimeError(f"el comando fallo con codigo {e.returncode}") from e


def guardar_movimiento_en_archivo(num, fecha, nombre, monto, tipo=TIPO_CHEQUE, conciliacion=ESTADO_TRANSITO):
    with open(ARCHIVO_CHEQUES, mode="a", newline="", encoding="utf-8") as f:
        escritor = csv.writer(f)
        escritor.writerow([num, fecha, nombre, formatear_monto(monto), "TRANSITO", tipo, conciliacion])


def guardar_en_archivo(num, fecha, nombre, monto):
    guardar_movimiento_en_archivo(num, fecha, nombre, monto, TIPO_CHEQUE, ESTADO_TRANSITO)
    print("💾 Datos guardados en el historial (CSV).")


def guardar_cheque_en_archivo(num, fecha, nombre, monto):
    guardar_movimiento_en_archivo(num, fecha, nombre, monto, TIPO_CHEQUE, ESTADO_TRANSITO)


def guardar_cheques_registrados_en_archivo(df):
    columnas = ["Num", "Fecha", "Nombre", "Monto", "Estado", "Tipo", "Conciliacion"]
    if df.empty:
        return

    df = df.copy()
    for columna in columnas:
        if columna not in df.columns:
            df[columna] = ""

    df[columnas].to_csv(ARCHIVO_CHEQUES, index=False, header=False)


def guardar_depositos_registrados_en_archivo(df):
    columnas = ["Fecha", "Descripcion", "Monto", "Conciliacion"]
    if df.empty:
        return

    df = df.copy()
    for columna in columnas:
        if columna not in df.columns:
            df[columna] = ""

    df[columnas].to_csv(ARCHIVO_DEPOSITOS, index=False, header=False)


def emitir_cheque_datos(num_cheque, nombre, monto, fecha=None):
    num_cheque = normalizar_numero_cheque(num_cheque)
    if not num_cheque:
        raise ErrorOperacion("⚠️ Error: el número de cheque debe ser mayor que cero.")

    nombre = str(nombre or "").strip()
    if not nombre:
        raise ErrorOperacion("⚠️ Error: el campo no puede quedar vacío.")

    monto = convertir_monto(monto)
    if monto is None:
        raise ErrorOperacion("⚠️ Error: Solo usar números y punto decimal.")
    if monto <= 0:
        raise ErrorOperacion("⚠️ Error: El monto debe ser mayor que cero.")

    if cheque_ya_registrado(num_cheque):
        raise ErrorOperacion(f"⚠️ El cheque {num_cheque} ya existe en el historial.")

    fecha = fecha or datetime.now().strftime("%Y-%m-%d")
    nombre = nombre.upper()

    imprimir_cheque_pdf(num_cheque, fecha, nombre, monto)
    guardar_cheque_en_archivo(num_cheque, fecha, nombre, monto)

    return {
        "num": num_cheque,
        "fecha": fecha,
        "nombre": nombre,
        "monto": monto,
        "pdf": f"cheque_{num_cheque}.pdf",
        "mensaje": "✅ Cheque registrado y listo para imprimir con éxito.",
    }


def registrar_e_imprimir():
    print("\n--- EMISIÓN DE CHEQUE ---")

    monto = pedir_monto_positivo("Monto del cheque (ej. 1500.50): ")
    nombre = pedir_texto_no_vacio("Páguese a la orden de: ").upper()
    num_cheque = pedir_numero_cheque("Número de cheque: ")
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    if cheque_ya_registrado(num_cheque):
        print(f"⚠️ El cheque {num_cheque} ya existe en el historial.")
        return

    try:
        imprimir_cheque_pdf(num_cheque, fecha_actual, nombre, monto)
        guardar_en_archivo(num_cheque, fecha_actual, nombre, monto)
    except Exception as e:
        print(f"⚠️ No se pudo completar la emisión del cheque: {e}")
        return

    print("✅ Cheque registrado y listo para imprimir con éxito.")


def registrar_nota_debito_datos(referencia, descripcion, monto, fecha=None):
    referencia = str(referencia or "").strip().upper()
    if not referencia:
        raise ErrorOperacion("⚠️ Error: la referencia de la nota de débito no puede quedar vacía.")

    descripcion = str(descripcion or "").strip()
    if not descripcion:
        raise ErrorOperacion("⚠️ Error: el campo no puede quedar vacío.")

    monto = convertir_monto(monto)
    if monto is None:
        raise ErrorOperacion("⚠️ Error: Solo usar números y punto decimal.")
    if monto <= 0:
        raise ErrorOperacion("⚠️ Error: El monto debe ser mayor que cero.")

    if nota_debito_ya_registrada(referencia):
        raise ErrorOperacion(f"⚠️ La nota de débito {referencia} ya existe en el historial.")

    fecha = fecha or datetime.now().strftime("%Y-%m-%d")
    descripcion = descripcion.upper()
    guardar_movimiento_en_archivo(referencia, fecha, descripcion, monto, TIPO_NOTA_DEBITO, ESTADO_TRANSITO)

    return {
        "referencia": referencia,
        "fecha": fecha,
        "descripcion": descripcion,
        "monto": monto,
        "mensaje": "✅ Nota de débito registrada con éxito.",
    }


def registrar_nota_debito():
    print("\n--- REGISTRO DE NOTA DE DÉBITO ---")
    monto = pedir_monto_positivo("Monto de la nota de débito (ej. 1500.50): ")
    descripcion = pedir_texto_no_vacio("Descripción (ej. Pago de servicios): ").upper()
    referencia = pedir_texto_no_vacio("Número o referencia de nota de débito: ").upper()

    try:
        resultado = registrar_nota_debito_datos(referencia, descripcion, monto)
    except ErrorOperacion as e:
        print(e)
        return

    print(resultado["mensaje"])


def anular_cheque_numero(num_anular):
    if not os.path.exists(ARCHIVO_CHEQUES):
        raise ErrorOperacion("⚠️ No hay registro de cheques aún.")

    num_anular = normalizar_numero_cheque(num_anular)
    if not num_anular:
        raise ErrorOperacion("⚠️ Error: el número de cheque debe ser mayor que cero.")

    df = cargar_cheques_registrados()

    if df.empty:
        raise ErrorOperacion("⚠️ No hay registro de cheques aún.")

    coincidencias = df["Num_norm"].eq(num_anular)
    if not coincidencias.any():
        raise ErrorOperacion(f"⚠️ El cheque {num_anular} no existe en los registros.")

    df.loc[coincidencias, "Estado"] = "ANULADO"
    df.loc[coincidencias, "Conciliacion"] = ESTADO_ANULADO
    df[["Num", "Fecha", "Nombre", "Monto", "Estado", "Tipo", "Conciliacion"]].to_csv(
        ARCHIVO_CHEQUES,
        index=False,
        header=False,
    )

    cantidad = int(coincidencias.sum())
    if cantidad > 1:
        mensaje = f"⚠️ El número {num_anular} aparecía {cantidad} veces y quedó marcado como ANULADO."
    else:
        mensaje = f"🚫 ¡Hecho! El cheque {num_anular} ha sido marcado como ANULADO."

    return {"num": num_anular, "cantidad": cantidad, "mensaje": mensaje}


def anular_cheque():
    print("\n--- ANULAR CHEQUE ---")

    if not os.path.exists(ARCHIVO_CHEQUES):
        print("⚠️ No hay registro de cheques aún.")
        return

    num_anular = pedir_numero_cheque("Ingresa el número de cheque que deseas anular: ")
    try:
        resultado = anular_cheque_numero(num_anular)
    except ErrorOperacion as e:
        print(e)
        return
    print(resultado["mensaje"])


def obtener_conciliacion():
    if not os.path.exists(ARCHIVO_CHEQUES):
        raise ErrorOperacion("⚠️ Faltan archivos. Asegúrate de tener registros y estado de cuenta.")

    try:
        df_nuestro = cargar_cheques_registrados()
        df_banco = cargar_estado_cuenta_banco()
        df_depositos = cargar_depositos_registrados()
        df_nuestro["Conciliacion"] = ESTADO_TRANSITO
        if not df_depositos.empty:
            df_depositos["Conciliacion"] = ESTADO_TRANSITO

        cheques = []
        banco_usado = set()
        for idx_local, fila in df_nuestro.iterrows():
            num = fila.get("Num_norm")
            referencia = str(fila.get("Num", num or "")).strip()
            referencia_norm = normalizar_referencia(referencia)
            if not num and not referencia_norm:
                continue

            estado = str(fila.get("Estado", "")).upper()
            tipo = str(fila.get("Tipo", TIPO_CHEQUE)).upper()
            etiqueta = "Nota de débito" if tipo == TIPO_NOTA_DEBITO else "Cheque"
            identificador = referencia or num
            debitos = df_banco[df_banco["Tipo_movimiento"] == "DEBITO"].copy()

            if tipo == TIPO_NOTA_DEBITO:
                cobrado = debitos[debitos["Referencia_norm"].eq(referencia_norm)]
            else:
                cobrado = debitos[debitos["Num_norm"].eq(num)]

            monto_nuestro = fila["Monto_valor"]
            fecha_nuestra = fila.get("Fecha_dt")
            if cobrado.empty and monto_nuestro is not None:
                candidatos = debitos[debitos["Monto_valor"].eq(monto_nuestro)]
                if pd.notna(fecha_nuestra):
                    candidatos = candidatos[candidatos["Fecha_dt"].eq(fecha_nuestra)]
                cobrado = candidatos

            if estado == "ANULADO":
                if cobrado.empty:
                    mensaje = f"🚫 {etiqueta} {identificador} está ANULADO y no aparece cobrado en el banco."
                    resultado = "ANULADO"
                    df_nuestro.at[idx_local, "Conciliacion"] = ESTADO_ANULADO
                else:
                    banco_usado.add(cobrado.index[0])
                    mensaje = f"🚨 {etiqueta} {identificador} está ANULADO pero el banco sí lo cobró."
                    resultado = "ALERTA"
                    df_nuestro.at[idx_local, "Conciliacion"] = ESTADO_CONFLICTO
                cheques.append({"num": identificador, "estado": estado, "resultado": resultado, "mensaje": mensaje})
                continue

            if monto_nuestro is None:
                if not cobrado.empty:
                    banco_usado.add(cobrado.index[0])
                mensaje = f"⚠️ No pude comparar el monto de {etiqueta.lower()} {identificador} por datos inválidos."
                df_nuestro.at[idx_local, "Conciliacion"] = ESTADO_INVALIDO
                cheques.append(
                    {
                        "num": identificador,
                        "estado": estado,
                        "resultado": "INVALIDO",
                        "monto_nuestro": monto_nuestro,
                        "monto_banco": cobrado.iloc[0]["Monto_valor"] if not cobrado.empty else None,
                        "mensaje": mensaje,
                    }
                )
                continue

            if cobrado.empty:
                mensaje = f"⏳ {etiqueta} {identificador} en TRÁNSITO."
                df_nuestro.at[idx_local, "Conciliacion"] = ESTADO_TRANSITO
                cheques.append({"num": identificador, "estado": estado, "resultado": "TRANSITO", "mensaje": mensaje})
                continue

            monto_banco = cobrado.iloc[0]["Monto_valor"]
            banco_usado.add(cobrado.index[0])

            if monto_nuestro is None or monto_banco is None:
                mensaje = f"⚠️ No pude comparar el monto de {etiqueta.lower()} {identificador} por datos inválidos."
                resultado = "INVALIDO"
                df_nuestro.at[idx_local, "Conciliacion"] = ESTADO_INVALIDO
            elif monto_nuestro == monto_banco:
                mensaje = f"✅ {etiqueta} {identificador} cobrado perfectamente."
                resultado = "COBRADO"
                df_nuestro.at[idx_local, "Conciliacion"] = ESTADO_RECONCILIADO
            else:
                mensaje = (
                    f"⚠️ ¡Ojo! {etiqueta} {identificador} diferencia: Nuestro Q {formatear_monto(monto_nuestro)} | "
                    f"Banco Q {formatear_monto(monto_banco)}"
                )
                resultado = "DIFERENCIA"
                df_nuestro.at[idx_local, "Conciliacion"] = ESTADO_DIFERENCIA
            cheques.append(
                {
                    "num": identificador,
                    "estado": estado,
                    "resultado": resultado,
                    "monto_nuestro": monto_nuestro,
                    "monto_banco": monto_banco,
                    "mensaje": mensaje,
                }
            )

        depositos = []
        creditos = df_banco[df_banco["Tipo_movimiento"] == "CREDITO"].copy()
        for idx_local, fila in df_depositos.iterrows():
            monto_nuestro = fila["Monto_valor"]
            fecha_nuestra = fila.get("Fecha_dt")
            descripcion = str(fila.get("Descripcion", "")).strip()
            if monto_nuestro is None:
                df_depositos.at[idx_local, "Conciliacion"] = ESTADO_INVALIDO
                continue

            disponibles = creditos[~creditos.index.isin(banco_usado)]
            acreditado = disponibles[disponibles["Monto_valor"].eq(monto_nuestro)]
            if pd.notna(fecha_nuestra):
                acreditado = acreditado[acreditado["Fecha_dt"].eq(fecha_nuestra)]

            if acreditado.empty:
                mensaje = f"⏳ Depósito {descripcion} por Q {formatear_monto(monto_nuestro)} pendiente de aparecer en banco."
                df_depositos.at[idx_local, "Conciliacion"] = ESTADO_TRANSITO
                depositos.append({"descripcion": descripcion, "resultado": "TRANSITO", "mensaje": mensaje})
                continue

            banco_usado.add(acreditado.index[0])
            mensaje = f"✅ Depósito {descripcion} por Q {formatear_monto(monto_nuestro)} acreditado en banco."
            df_depositos.at[idx_local, "Conciliacion"] = ESTADO_RECONCILIADO
            depositos.append({"descripcion": descripcion, "resultado": "ACREDITADO", "mensaje": mensaje})

        no_registrados = []
        for indice, fila_banco in df_banco.iterrows():
            if indice in banco_usado:
                continue

            monto_banco = fila_banco["Monto_valor"]
            monto_texto = formatear_monto(monto_banco) if monto_banco is not None else "N/D"
            num_bco = fila_banco.get("Num_norm")
            referencia = fila_banco.get("Referencia") or num_bco or "S/R"
            tt = str(fila_banco.get("TT", "")).strip()
            detalle = f"{tt} {referencia}".strip()

            if fila_banco["Tipo_movimiento"] == "DEBITO":
                mensaje = (
                    f"❓ Cheque {num_bco} por Q {monto_texto} cobrado por el banco, "
                    "pero NO está en nuestro sistema."
                ) if num_bco else (
                    f"❓ Cargo {detalle} por Q {monto_texto} aparece en banco, "
                    "pero NO está en nuestro sistema."
                )
            else:
                mensaje = (
                    f"❓ Crédito {detalle} por Q {monto_texto} aparece en banco, "
                    "pero NO está en depósitos o notas de crédito del sistema."
                )
            no_registrados.append({"num": referencia, "monto": monto_banco, "mensaje": mensaje})

        guardar_cheques_registrados_en_archivo(df_nuestro)
        guardar_depositos_registrados_en_archivo(df_depositos)

        return {"cheques": cheques, "depositos": depositos, "no_registrados": no_registrados}

    except ErrorOperacion:
        raise
    except Exception as e:
        raise ErrorOperacion(f"⚠️ Ocurrió un error leyendo los archivos: {e}") from e


def conciliar_cuentas():
    print("\n--- CONCILIACIÓN BANCARIA ---")

    try:
        resultado = obtener_conciliacion()
    except ErrorOperacion as e:
        print(e)
        return

    print("\n--- Resultados de Egresos Emitidos ---")
    for fila in resultado["cheques"]:
        print(fila["mensaje"])

    print("\n--- Resultados de Depósitos Registrados ---")
    for fila in resultado["depositos"]:
        print(fila["mensaje"])

    print("\n--- Movimientos del Banco No Registrados por Nosotros ---")
    for fila in resultado["no_registrados"]:
        print(fila["mensaje"])


def limpiar_texto_pdf(texto):
    reemplazos = {
        "✅": "[OK]",
        "⚠️": "[ALERTA]",
        "⚠": "[ALERTA]",
        "🚫": "[ANULADO]",
        "🚨": "[ALERTA]",
        "⏳": "[PENDIENTE]",
        "❓": "[SIN REGISTRO]",
    }
    texto = str(texto or "")
    for origen, destino in reemplazos.items():
        texto = texto.replace(origen, destino)
    return texto.encode("latin-1", errors="ignore").decode("latin-1")


def dividir_texto_pdf(texto, ancho=100):
    palabras = limpiar_texto_pdf(texto).split()
    if not palabras:
        return [""]

    lineas = []
    actual = ""
    for palabra in palabras:
        candidata = f"{actual} {palabra}".strip()
        if len(candidata) <= ancho:
            actual = candidata
            continue

        if actual:
            lineas.append(actual)
        actual = palabra

    if actual:
        lineas.append(actual)
    return lineas


def nombre_pdf_conciliacion(fecha=None):
    fecha = fecha or datetime.now()
    fecha = pd.Timestamp(fecha).to_pydatetime()
    return f"conciliacion_{fecha.strftime('%Y%m%d_%H%M%S')}.pdf"


def generar_pdf_conciliacion(resultado, fecha=None, nombre_pdf=None):
    fecha = fecha or datetime.now()
    fecha_dt = pd.Timestamp(fecha).to_pydatetime()
    nombre_pdf = nombre_pdf or nombre_pdf_conciliacion(fecha_dt)

    pdf = canvas.Canvas(nombre_pdf, pagesize=letter)
    ancho_pagina, alto_pagina = letter
    margen_x = 2 * cm
    margen_inferior = 2 * cm
    y = alto_pagina - 2 * cm

    def nueva_pagina():
        nonlocal y
        pdf.showPage()
        y = alto_pagina - 2 * cm

    def escribir(texto, fuente="Helvetica", tamano=10, salto=0.45 * cm, indent=0, ancho=100):
        nonlocal y
        pdf.setFont(fuente, tamano)
        for linea in dividir_texto_pdf(texto, ancho=ancho):
            if y < margen_inferior:
                nueva_pagina()
                pdf.setFont(fuente, tamano)
            pdf.drawString(margen_x + indent, y, linea)
            y -= salto

    def espacio(alto=0.25 * cm):
        nonlocal y
        y -= alto
        if y < margen_inferior:
            nueva_pagina()

    def seccion(titulo, filas, mensaje_vacio):
        escribir(titulo, "Helvetica-Bold", 12, salto=0.55 * cm)
        if filas:
            for fila in filas:
                escribir(f"- {fila['mensaje']}", indent=0.35 * cm, ancho=94)
        else:
            escribir(mensaje_vacio, indent=0.35 * cm)
        espacio()

    escribir("REPORTE DE CONCILIACION BANCARIA", "Helvetica-Bold", 15, salto=0.65 * cm)
    escribir(f"Generado: {fecha_dt.strftime('%Y-%m-%d %H:%M')}", "Helvetica", 10)
    espacio()
    escribir("Resumen", "Helvetica-Bold", 12, salto=0.55 * cm)
    escribir(f"Egresos revisados: {len(resultado.get('cheques', []))}", indent=0.35 * cm)
    escribir(f"Depositos revisados: {len(resultado.get('depositos', []))}", indent=0.35 * cm)
    escribir(f"Movimientos no registrados: {len(resultado.get('no_registrados', []))}", indent=0.35 * cm)
    espacio()

    seccion(
        "Resultados de egresos emitidos",
        resultado.get("cheques", []),
        "Sin egresos registrados para conciliar.",
    )
    seccion(
        "Resultados de depositos registrados",
        resultado.get("depositos", []),
        "Sin depositos registrados para conciliar.",
    )
    seccion(
        "Movimientos del banco no registrados por nosotros",
        resultado.get("no_registrados", []),
        "Sin movimientos pendientes de registrar.",
    )

    pdf.save()
    return nombre_pdf


def imprimir_conciliacion():
    print("\n--- IMPRIMIR CONCILIACIÓN BANCARIA ---")

    try:
        resultado = obtener_conciliacion()
        nombre_pdf = generar_pdf_conciliacion(resultado)
    except ErrorOperacion as e:
        print(e)
        return

    print(f"📄 Reporte de conciliación generado: {nombre_pdf}")
    print("Puedes abrir ese PDF e imprimirlo para el archivo de contabilidad.")


def clear_ide_terminal():
    """Pushes old text out of view using standard line breaks."""
    print("\n" * 45)


def obtener_reporte_movimientos(fecha=None):
    fecha = fecha or datetime.now().date()
    periodo_actual = pd.Timestamp(fecha).to_period("M")
    total_cheques = Decimal("0.00")
    total_depositos = Decimal("0.00")

    df_cheques = cargar_cheques_registrados()
    df_cheques = df_cheques[df_cheques["Monto_valor"].notna() & df_cheques["Fecha_dt"].notna()].copy()
    df_cheques_mes = df_cheques[df_cheques["Fecha_dt"].dt.to_period("M") == periodo_actual].copy()

    if not df_cheques_mes.empty:
        df_cheques_mes["Monto"] = df_cheques_mes["Monto_valor"].map(formatear_monto)
        total_cheques = sum(
            (fila["Monto_valor"] for _, fila in df_cheques_mes.iterrows() if str(fila["Estado"]).upper() != "ANULADO"),
            Decimal("0.00"),
        )
    else:
        df_cheques_mes = pd.DataFrame(columns=["Num", "Fecha", "Nombre", "Monto", "Estado", "Tipo"])

    df_depositos = cargar_depositos_registrados()
    df_depositos = df_depositos[df_depositos["Monto_valor"].notna() & df_depositos["Fecha_dt"].notna()].copy()
    df_depositos_mes = df_depositos[df_depositos["Fecha_dt"].dt.to_period("M") == periodo_actual].copy()

    if not df_depositos_mes.empty:
        df_depositos_mes["Monto"] = df_depositos_mes["Monto_valor"].map(formatear_monto)
        total_depositos = sum(
            (fila["Monto_valor"] for _, fila in df_depositos_mes.iterrows()),
            Decimal("0.00"),
        )
    else:
        df_depositos_mes = pd.DataFrame(columns=["Fecha", "Descripcion", "Monto"])

    saldo = total_depositos - total_cheques

    return {
        "periodo": str(periodo_actual),
        "cheques": df_cheques_mes,
        "depositos": df_depositos_mes,
        "total_cheques": total_cheques,
        "total_depositos": total_depositos,
        "saldo": saldo,
    }


def reporte_movimientos():
    print("\n" + "=" * 40)
    print(" 📊 CORTE DE CAJA MENSUAL 📊")
    print("=" * 40)

    reporte = obtener_reporte_movimientos()
    df_cheques_mes = reporte["cheques"]

    if not df_cheques_mes.empty:
        print("\n--- EGRESOS DEL MES ---")
        print(df_cheques_mes[["Num", "Fecha", "Nombre", "Monto", "Estado", "Tipo"]].to_string(index=False))
    else:
        print("\n--- EGRESOS DEL MES ---")
        print("No hay cheques ni notas de débito registrados en el mes actual.")

    df_depositos_mes = reporte["depositos"]

    if not df_depositos_mes.empty:
        print("\n--- DEPÓSITOS RECIBIDOS DEL MES ---")
        print(df_depositos_mes[["Fecha", "Descripcion", "Monto"]].to_string(index=False))
    else:
        print("\n--- DEPÓSITOS RECIBIDOS DEL MES ---")
        print("No hay depósitos registrados en el mes actual.")

    print("\n" + "-" * 40)
    print(f"📈 TOTAL INGRESOS (Depósitos): Q {formatear_monto(reporte['total_depositos'])}")
    print(f"📉 TOTAL EGRESOS (Cheques y notas de débito): Q {formatear_monto(reporte['total_cheques'])}")
    print("-" * 40)
    print(f"💵 SALDO EN BÓVEDA:          Q {formatear_monto(reporte['saldo'])}")
    print("-" * 40)

    input("\nPresiona ENTER para volver al menú...")


def mostrar_menu():
    print("\n" + "=" * 40)
    print(" 💼 SISTEMA DE CONTROL BANCARIO 💼")
    print("=" * 40)
    print("1. Emitir nuevo cheque")
    print("2. Registrar nota de débito")
    print("3. Registrar un depósito")
    print("4. Conciliar banco")
    print("5. Anular un cheque con error")
    print("6. Ver corte de caja (Reporte)")
    print("7. Imprimir conciliación")
    print("8. Salir")
    print("=" * 40)


def main():
    while True:
        mostrar_menu()
        opcion = input("Elige una opción (1-8): ").strip()

        if opcion == "1":
            registrar_e_imprimir()
        elif opcion == "2":
            registrar_nota_debito()
        elif opcion == "3":
            registrar_deposito()
        elif opcion == "4":
            conciliar_cuentas()
        elif opcion == "5":
            anular_cheque()
        elif opcion == "6":
            reporte_movimientos()
        elif opcion == "7":
            imprimir_conciliacion()
        elif opcion == "8":
            print("\nCerrando la bóveda... ¡Buenas noches y éxito en los negocios, Javier!")
            sys.exit()
        else:
            print("⚠️ Opción inválida. Intenta un número del 1 al 8.")


if __name__ == "__main__":  # pragma: no cover
    main()
